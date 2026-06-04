from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.rules.models import (
    AssociationTransaction,
    AssociationTransactionItem,
    FoodSupplementSynergyRule,
    SafetyConstraint,
    SupplementCategory,
    SupplementNormalization,
)
from apps.rules.services import canonical_key, normalize_rule_item, split_keywords


REQUIRED_SHEETS = {
    "Supplement_Categories",
    "Normalization_105",
    "Synergy_Seed_Rules",
    "Safety_Constraints",
    "Transactions_Wide",
    "Transactions_Long",
    "Apriori_OneHot",
}

REQUIRED_COLUMNS = {
    "Supplement_Categories": {"category", "canonical_item", "keywords", "main_nutrient", "source_url"},
    "Normalization_105": {"original_supplement_name", "normalized_category", "primary_keyword", "main_nutrient", "notes", "source_url"},
    "Synergy_Seed_Rules": {"rule_seed_id", "supplement_category", "supplement_item", "food", "food_item", "nutrient_relation", "association_type", "reason", "seed_weight", "source_url"},
    "Safety_Constraints": {"supplement_category", "avoid_or_review_item", "constraint_type", "reason", "how_to_use", "source_url"},
    "Transactions_Wide": {"transaction_id", "items_pipe"},
    "Transactions_Long": {"transaction_id", "item_type", "item_value", "item"},
    "Apriori_OneHot": {"transaction_id"},
}


class Command(BaseCommand):
    help = "Import I-NutriGuide association-rule Excel dataset."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")

    def handle(self, *args, **options):
        path = self._resolve_path(options["xlsx_path"])
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required. Install backend requirements before importing .xlsx files.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        self._validate_workbook(workbook)
        summary = {"imported": 0, "updated": 0, "skipped": 0, "errors": 0}

        with transaction.atomic():
            self._import_categories(workbook, summary)
            self._import_normalizations(workbook, summary)
            self._import_synergy_rules(workbook, summary)
            self._import_safety_constraints(workbook, summary)
            self._import_transactions_wide(workbook, summary)
            self._import_transactions_long(workbook, summary)
            self._import_one_hot(workbook, summary)

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: imported={summary['imported']}, updated={summary['updated']}, "
                f"skipped={summary['skipped']}, errors={summary['errors']}"
            )
        )

    def _resolve_path(self, raw_path):
        candidates = [
            Path(raw_path),
            Path.cwd() / raw_path,
            Path.cwd() / "data" / Path(raw_path).name,
        ]
        for path in candidates:
            if path.exists():
                return path
        raise CommandError(f"Excel file not found: {raw_path}")

    def _validate_workbook(self, workbook):
        missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
        if missing_sheets:
            raise CommandError(f"Missing required sheet(s): {', '.join(missing_sheets)}")
        for sheet_name, columns in REQUIRED_COLUMNS.items():
            header = set(self._headers(workbook[sheet_name]))
            missing = sorted(columns - header)
            if missing:
                raise CommandError(f"{sheet_name} missing required column(s): {', '.join(missing)}")

    def _headers(self, sheet):
        return [self._text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if self._text(value)]

    def _rows(self, sheet):
        iterator = sheet.iter_rows(values_only=True)
        headers = [self._text(value) for value in next(iterator)]
        for values in iterator:
            row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
            if any(self._text(value) for value in row.values()):
                yield row

    def _import_categories(self, workbook, summary):
        for row in self._rows(workbook["Supplement_Categories"]):
            category = self._text(row.get("category"))
            canonical_item = self._text(row.get("canonical_item")) or canonical_key(category)
            if not category or not canonical_item:
                summary["skipped"] += 1
                continue
            _obj, created = SupplementCategory.objects.update_or_create(
                canonical_item=canonical_key(canonical_item),
                defaults={
                    "category": category,
                    "keywords": split_keywords(row.get("keywords")),
                    "main_nutrient": self._text(row.get("main_nutrient")),
                    "source_url": self._text(row.get("source_url")),
                    "is_active": True,
                },
            )
            summary["imported" if created else "updated"] += 1

    def _import_normalizations(self, workbook, summary):
        for row in self._rows(workbook["Normalization_105"]):
            original_name = self._text(row.get("original_supplement_name"))
            normalized_category = self._text(row.get("normalized_category"))
            if not original_name or not normalized_category:
                summary["skipped"] += 1
                continue
            category = self._category_for(normalized_category)
            _obj, created = SupplementNormalization.objects.update_or_create(
                original_supplement_slug=canonical_key(original_name),
                defaults={
                    "original_supplement_name": original_name,
                    "category": category,
                    "normalized_category": normalized_category,
                    "primary_keyword": self._text(row.get("primary_keyword")),
                    "main_nutrient": self._text(row.get("main_nutrient")),
                    "notes": self._text(row.get("notes")),
                    "source_url": self._text(row.get("source_url")),
                    "is_active": True,
                },
            )
            summary["imported" if created else "updated"] += 1

    def _import_synergy_rules(self, workbook, summary):
        for row in self._rows(workbook["Synergy_Seed_Rules"]):
            rule_seed_id = self._text(row.get("rule_seed_id"))
            supplement_category = self._text(row.get("supplement_category"))
            food_item = normalize_rule_item(row.get("food_item"))
            supplement_item = normalize_rule_item(row.get("supplement_item"))
            if not rule_seed_id or not supplement_category or not food_item or not supplement_item:
                summary["skipped"] += 1
                continue
            _obj, created = FoodSupplementSynergyRule.objects.update_or_create(
                rule_seed_id=rule_seed_id,
                defaults={
                    "supplement_category": self._category_for(supplement_category),
                    "supplement_category_name": supplement_category,
                    "supplement_item": supplement_item,
                    "food": self._text(row.get("food")),
                    "food_item": food_item,
                    "nutrient_relation": self._text(row.get("nutrient_relation")),
                    "association_type": canonical_key(row.get("association_type")) or "positive",
                    "reason": self._text(row.get("reason")),
                    "seed_weight": self._float(row.get("seed_weight"), 0.75),
                    "source_url": self._text(row.get("source_url")),
                    "is_active": True,
                },
            )
            summary["imported" if created else "updated"] += 1

    def _import_safety_constraints(self, workbook, summary):
        for row in self._rows(workbook["Safety_Constraints"]):
            supplement_category = self._text(row.get("supplement_category"))
            avoid_item = self._text(row.get("avoid_or_review_item"))
            constraint_type = canonical_key(row.get("constraint_type"))
            if not supplement_category or not avoid_item or not constraint_type:
                summary["skipped"] += 1
                continue
            _obj, created = SafetyConstraint.objects.update_or_create(
                supplement_category_name=supplement_category,
                avoid_or_review_item=avoid_item,
                constraint_type=constraint_type,
                defaults={
                    "supplement_category": self._category_for(supplement_category),
                    "reason": self._text(row.get("reason")),
                    "how_to_use": self._text(row.get("how_to_use")),
                    "source_url": self._text(row.get("source_url")),
                    "is_active": True,
                },
            )
            summary["imported" if created else "updated"] += 1

    def _import_transactions_wide(self, workbook, summary):
        for row in self._rows(workbook["Transactions_Wide"]):
            transaction_id = self._text(row.get("transaction_id"))
            if not transaction_id:
                summary["skipped"] += 1
                continue
            raw_payload = {key: self._text(value) for key, value in row.items() if self._text(value)}
            _obj, created = AssociationTransaction.objects.update_or_create(
                transaction_id=transaction_id,
                defaults={"source": "association_excel", "raw_payload": raw_payload},
            )
            summary["imported" if created else "updated"] += 1

    def _import_transactions_long(self, workbook, summary):
        for row in self._rows(workbook["Transactions_Long"]):
            transaction_id = self._text(row.get("transaction_id"))
            item = normalize_rule_item(row.get("item"))
            if not transaction_id or not item:
                summary["skipped"] += 1
                continue
            transaction_obj, transaction_created = AssociationTransaction.objects.get_or_create(
                transaction_id=transaction_id,
                defaults={"source": "association_excel"},
            )
            summary["imported" if transaction_created else "updated"] += 1
            _obj, created = AssociationTransactionItem.objects.update_or_create(
                transaction=transaction_obj,
                item=item,
                defaults={
                    "item_type": canonical_key(row.get("item_type")),
                    "item_value": canonical_key(row.get("item_value")),
                },
            )
            summary["imported" if created else "updated"] += 1

    def _import_one_hot(self, workbook, summary):
        sheet = workbook["Apriori_OneHot"]
        rows = sheet.iter_rows(values_only=True)
        headers = [self._text(value) for value in next(rows)]
        item_headers = [normalize_rule_item(header) for header in headers[1:]]
        for values in rows:
            transaction_id = self._text(values[0] if values else "")
            if not transaction_id:
                summary["skipped"] += 1
                continue
            active_items = [
                item
                for index, item in enumerate(item_headers, start=1)
                if item and self._truthy(values[index] if index < len(values) else None)
            ]
            _obj, created = AssociationTransaction.objects.update_or_create(
                transaction_id=transaction_id,
                defaults={"source": "association_excel", "one_hot_items": active_items},
            )
            summary["imported" if created else "updated"] += 1

    def _category_for(self, category):
        if not category:
            return None
        return SupplementCategory.objects.filter(category__iexact=category).first()

    def _text(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _float(self, value, default):
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def _truthy(self, value):
        return self._text(value).lower() in {"1", "true", "yes", "y"}
